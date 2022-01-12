import os
import sys
import webbrowser
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import QDate
from PyQt5 import uic
from openpyxl import load_workbook
import subprocess
import shutil
import pyautogui
import pyperclip
import time

# os.environ['QT_MAC_WANTS_LAYER'] = '1' #필수

form_class = uic.loadUiType("/Users/myeongho/Codes/메이킷코드/Personal_Project/PersonalProject.ui")[0]

class MyWindow(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle("메이킷코드랩")

        self.pushButton.clicked.connect(self.open_webbrowser_calender)
        self.pushButton_2.clicked.connect(self.open_webbrowser_notes)
        self.pushButton_3.clicked.connect(self.open_webbrowser_correction)
        self.pushButton_4.clicked.connect(self.open_webbrowser_ACAweb)
        self.pushButton_6.clicked.connect(self.NasUpLoad)
        self.pushButton_12.clicked.connect(self.OpenMakIT)
        self.pushButton_13.clicked.connect(self.Class)
        self.pushButton_14.clicked.connect(self.Daily_Report_KakaoTalk)
        self.pushButton_15.clicked.connect(self.Change_File_Name)
        self.pushButton_16.clicked.connect(self.Send_KakaoTalk)
        self.pushButton_18.clicked.connect(self.Create_readme)
        self.pushButton_19.clicked.connect(self.Open_Readme)
        self.pushButton_20.clicked.connect(self.Readme)
        self.pushButton_21.clicked.connect(self.open_kakaoTalk)
        self.pushButton_22.clicked.connect(self.autoxl)
        self.inquiry() #statusBar에 시간 출력하기

    

    def OpenFolder(self, Path):
        file_to_show = Path
        subprocess.call(["open", "-R", file_to_show])

    def CopySth(self, Text): # 클립보드에 
        cb = QApplication.clipboard()
        cb.clear(mode=cb.Clipboard)
        cb.setText(Text, mode=cb.Clipboard)

    def Hot_Key(self, key, press):
        pyautogui.keyDown(key)
        pyautogui.press(press)
        pyautogui.keyUp(key)

    def open_webbrowser_calender(self):
        webbrowser.open('https://calendar.google.com/calendar/u/1/r?tab=rc')
    
    def open_webbrowser_notes(self):
        webbrowser.open('https://docs.google.com/spreadsheets/d/1nJQXdCRSKQGuwczvrQ8_aScwIqU9SfYiHfhy0Lk2R88/edit#gid=1648232156')
        
    def open_webbrowser_correction(self):
        webbrowser.open('https://speller.cs.pusan.ac.kr')

    def open_webbrowser_ACAweb(self):
        webbrowser.open('https://t.aca2000.co.kr/Account/Login?ReturnUrl=%2F')
    
    def Daily_Report_KakaoTalk(self): #데일리보고방 카톡
        str_date = cur_date.toString("MM.dd(ddd)")
        os.system("open /Applications/KakaoTalk.app")
        time.sleep(0.3)
        self.Hot_Key('command', '2')
        time.sleep(0.3)
        self.Hot_Key('command', 'f')
        time.sleep(0.1)
        pyperclip.copy('데일리보고방')
        self.Hot_Key('command', 'v')
        time.sleep(0.3)
        pyautogui.press('down')
        pyautogui.press('return')
        time.sleep(0.3)
        pyperclip.copy('[서명호 연구원 데일리 보고]\n\n-금일 '+str_date+'데일리 보고 및 수업 진도 관리 파일 첨부합니다.')
        self.Hot_Key("command", 'v')
        pyautogui.press("return")



        file_list = ['/Users/myeongho/Codes/메이킷코드/보고파일 3개/데일리보고','/Users/myeongho/Codes/메이킷코드/보고파일 3개/진도표']
        for i in file_list:

            self.Hot_Key("command", 'o')
            time.sleep(0.2)
            pyperclip.copy(i)
            pyautogui.keyDown("command");pyautogui.keyDown("shift")
            pyautogui.press("g")
            pyautogui.keyUp("command");pyautogui.keyUp("shift")
            self.Hot_Key("command", 'v');time.sleep(0.1)
            pyautogui.press('return')
            time.sleep(0.2)
            pyautogui.press('right')
            time.sleep(0.1)
            pyautogui.press('return')
            time.sleep(0.5)

    def OpenMakIT(self):
        self.OpenFolder("/Users/myeongho/Codes/메이킷코드/보고파일 3개")
        
    def Class(self):
        self.OpenFolder("/Users/myeongho/Library/Group Containers/G69SCX94XU.duck/Library/Application Support/duck/Volumes/makitedu.synology.me – WebDAV HTTPS/메이킷코드랩_파트/1.연구원/5.서명호연구원")

    def Change_File_Name(self):
        cur_date = QDate.currentDate()
        str_date = cur_date.toString("yyyy.MM.dd")
        path = '/Users/myeongho/Codes/메이킷코드/보고파일 3개'

        file_path = [path+'/진도표', path+'/데일리보고', path+'/카카오톡피드백']

        for i in range(0, len(file_path)):
            file_name = file_path[i]
            file_list = os.listdir(file_name)
            src = os.path.join(file_name, file_list[0])
            if(i == 0):
                dst = '[서명호 연구원]진도표'+str_date+'.xlsx'
            elif(i == 1):
                dst = '서명호 연구원_데일리보고'+str_date+'.pptx'
            elif(i==2):
                dst = '서명호연구원'+str_date+'.xlsx'

            dst = os.path.join(file_name, dst)
            os.rename(src,dst)

    
    def Create_readme(self, f): # 리드미 작성
        cur_date = QDate.currentDate()
        date_1 = cur_date.toString("M월")
        date_2 = cur_date.toString("M월dd일(ddd)")

        f = open ("/Users/myeongho/Library/Group Containers/G69SCX94XU.duck/Library/Application Support/duck/Volumes/makitedu.synology.me – WebDAV HTTPS/메이킷코드랩_파트/2.수업결과보고_대치센터/수업자료/"+date_1+'/'+date_2+'/readme.txt', 'w', encoding = "UTF8")

        Readme_List = [self.plainTextEdit_5.toPlainText(), self.plainTextEdit_6.toPlainText(), self.plainTextEdit_7.toPlainText(), self.plainTextEdit.toPlainText()]
        for i in range(len(Readme_List)):
            if (Readme_List[i] == ''):
                Readme_List[i] = "없습니다."
        
        
        text ="지난숙제: "+Readme_List[0]+"\n\n결석학생: "+Readme_List[1]+"\n\n진도: "+Readme_List[2]+"\n\n이번 숙제: "+Readme_List[3]

        try:
            f.write(text)
            f.close()
        except:
            f.write("ERROR")
            f.close()

    def Open_Readme(self):
        cur_date = QDate.currentDate()
        date_1 = cur_date.toString("M월")
        date_2 = cur_date.toString("M월dd일(ddd)")
        self.OpenFolder("/Users/myeongho/Library/Group Containers/G69SCX94XU.duck/Library/Application Support/duck/Volumes/makitedu.synology.me – WebDAV HTTPS/메이킷코드랩_파트/2.수업결과보고_대치센터/수업자료/"+date_1+"/"+date_2+'/')

    
    def open_kakaoTalk(self):
        self.OpenFolder("/Users/myeongho/Codes/메이킷코드/KAKAOTALK")

    def autoxl(self):#엑셀 자동 입력하기
        

        first_line = '“안녕하세요?\n메이킷코드랩 코딩학원입니다.\n\n'
        end_line = '\n\n메이킷코드랩 홈페이지 http://makitcodelab.com\n송도센터 032-833-0046\n대치센터 02-6243-5000"'

        homework_text = "(숙제) "+self.plainTextEdit_4.toPlainText()+"\n\n"
        class_text = self.plainTextEdit_2.toPlainText()
        student_text = self.plainTextEdit_3.toPlainText()


        cur_date = QDate.currentDate()
        str_date = cur_date.toString("yyyy년 MM월 dd일 dddd")
        str_date2 = cur_date.toString("MM.dd(ddd) ")
        file_name = str_date2+self.lineEdit.text()
        

        path = '/Users/myeongho/Codes/메이킷코드/KAKAOTALK/'+str_date2
        try:
            if not os.path.exists(path):
                os.makedirs(path)
        except OSError:
            print ('Error: Creating directory. ' +  path)

        if (self.lineEdit.text()== ""):
            file_name = str_date2 + '(임시저장)'

        file_list = os.listdir('/Users/myeongho/Codes/메이킷코드/보고파일 3개/카카오톡피드백/')
        wb = load_workbook('/Users/myeongho/Codes/메이킷코드/보고파일 3개/카카오톡피드백/'+file_list[0])
        ws = wb.active
        
        f = open(path+'/'+file_name+'.txt', 'w', encoding= 'UTF8')
        try:
            i = 2
            

            while ws.cell(row = i, column = 2).value !=None:
                i+=2
            

            if(ws.cell(row = i-2, column = 2).value == str_date):
                i-=2
                
            else:
                ws.cell(row = i, column =2, value = str_date)
            

            j =3
            
            while True:
                
                if(ws.cell(row = i+1, column = j).value == None):
                    if (self.checkBox.isChecked()==False):
                        msg = first_line+class_text+"\n\n"+student_text+end_line
                        name = self.lineEdit.text()
                        ws.cell (row = i+1, column= j , value = msg)
                        ws.cell (row = i, column=j, value= name)
                        f.write(msg)
                        f.close()
                    else:
                        msg = first_line + homework_text+class_text+'\n\n'+student_text+end_line
                        ws.cell (row = i+1, column= j , value = msg)
                        name = self.lineEdit.text()
                        ws.cell(row = i, column = j, value= name)
                        f.write(msg)
                        f.close()
                    break
                else:
                    j+=1
                    continue
                
        except:
            print("ERROR2")

        self.lineEdit.setText('')
        wb.save('/Users/myeongho/Codes/메이킷코드/보고파일 3개/카카오톡피드백/'+file_list[0])

    def NasUpLoad(self):
        path = '/Users/myeongho/Codes/메이킷코드/보고파일 3개'

        file_path = [path+'/진도표', path+'/데일리보고', path+'/카카오톡피드백']
        destination = ['/Users/myeongho/Library/Group Containers/G69SCX94XU.duck/Library/Application Support/duck/Volumes/makitedu.synology.me – WebDAV HTTPS/메이킷코드랩_파트/2.수업결과보고_대치센터/진도표/서명호_연구원', '/Users/myeongho/Library/Group Containers/G69SCX94XU.duck/Library/Application Support/duck/Volumes/makitedu.synology.me – WebDAV HTTPS/메이킷코드랩_파트/2.수업결과보고_대치센터/데일리_보고서/서명호_연구원', '/Users/myeongho/Library/Group Containers/G69SCX94XU.duck/Library/Application Support/duck/Volumes/makitedu.synology.me – WebDAV HTTPS/메이킷코드랩_파트/2.수업결과보고_대치센터/카톡_피드백(학부모)/서명호_연구원']


        for i in range(0 , len(file_path)):
            file_list = os.listdir(file_path[i])
            file_path[i] = file_path[i]+'/'+file_list[0]
        for i in range(3):
            shutil.copy(file_path[i], destination[i])

    def Send_KakaoTalk(self):
        cur_date = QDate.currentDate()
        str_date = cur_date.toString("yyyy년 MM월 dd일 dddd")

        file_list = os.listdir('/Users/myeongho/Codes/메이킷코드/보고파일 3개/카카오톡피드백/')
        wb = load_workbook('/Users/myeongho/Codes/메이킷코드/보고파일 3개/카카오톡피드백/'+file_list[0])
        ws = wb.active

        KakaoTalk = []
        i = 2
        while ws.cell(row = i, column = 2).value != str_date:
            i+=2
            if(i>=1000000):
                break
        
        j = 3
        while ws.cell(row= i+1, column = j).value != None:

            KakaoTalk.append(ws.cell(row= i+1, column = j).value)
            j+=1

        os.system("open /Applications/KakaoTalk.app")
        time.sleep(0.3)
        self.Hot_Key("command", '2')
        time.sleep(0.3)
        self.Hot_Key("command", 'f')
        time.sleep(0.3)
        pyperclip.copy('메이킷코드랩')
        self.Hot_Key("command", "v")
        time.sleep(0.3)
        pyautogui.press("down")
        time.sleep(0.3)
        pyautogui.press("return")
        
        for i in KakaoTalk:
            pyperclip.copy(i)
            self.Hot_Key("command", "v")
            time.sleep(0.1)
            pyautogui.press("return")
            

    def Readme(self):
        cur_date = QDate.currentDate()
        Month = cur_date.toString("M월")
        Day = cur_date.toString("M월dd일(ddd)")

        path = '/Users/myeongho/Library/Group Containers/G69SCX94XU.duck/Library/Application Support/duck/Volumes/makitedu.synology.me – WebDAV HTTPS/메이킷코드랩_파트/2.수업결과보고_대치센터/수업자료/'+Month+"/"+Day
        self.OpenFolder(path)
        if cur_date.toString("ddd") == "토":
            folder_name = [path+'/10시_스크_서명호연구원', path+'/13시_마이크로비트메이킹_서명호연구원', path+'/15시_파이썬5_서명호연구원']
            for path in folder_name:
                try:
                    if not os.path.exists(path):
                        os.makedirs(path)
                except OSError:
                    print ('Error: Creating directory. ' +  path)

        if cur_date.toString("ddd") == "일":
            folder_name = [path+'/13시_앱인벤터_서명호연구원', path+'/15시_파이썬1_서명호연구원']
            for path in folder_name:
                try:
                    if not os.path.exists(path):
                        os.makedirs(path)
                except OSError:
                    print ('Error: Creating directory. ' +  path)

        if cur_date.toString("ddd") == "화":
            folder_name = [path+'/13시_파이썬겨울방학특강_서명호연구원',path+'/16시20분_마이크로비트메이킹_서명호연구원']
            for path in folder_name:
                try:
                    if not os.path.exists(path):
                        os.makedirs(path)
                except OSError:
                    print ('Error: Creating directory. ' +  path)

        if cur_date.toString("ddd") == "수":
            folder_name = [path+'/13시_파이썬겨울방학특강_서명호연구원']
            for path in folder_name:
                try:
                    if not os.path.exists(path):
                        os.makedirs(path)
                except OSError:
                    print ('Error: Creating directory. ' +  path)



    def inquiry(self):
        cur_date = QDate.currentDate()
        str_date = cur_date.toString("yyyy년 MM월 dd일 dddd")
        self.statusBar().showMessage(str_date)


cur_date = QDate.currentDate()
app = QApplication(sys.argv)
window = MyWindow()
window.show()
app.exec_()

